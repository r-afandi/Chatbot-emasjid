import os
import json
import io
from openpyxl import load_workbook
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
import uvicorn

# Inisialisasi Aplikasi FastAPI
app = FastAPI(title="UI Data Knowledge Base")

# Nama file JSON untuk menyimpan data
JSON_FILE = "data_kategori_sementara.json"

def load_json_records(filepath: str) -> list:
    """Membaca data dari file JSON dengan aman."""
    if not os.path.exists(filepath) or os.path.getsize(filepath) == 0:
        return []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except json.JSONDecodeError:
        print(f"Peringatan: Gagal membaca {filepath}. File mungkin rusak. Memulai dengan list kosong.")
        return []


# Skema Data yang diterima dari Form (UI)
class LayananInput(BaseModel):
    nama: str = ""
    kategori: str
    pertanyaan: str
    jawaban: str

# Template HTML + CSS + JavaScript untuk UI
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Manajemen Data Chatbot</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f0f2f5; display: flex; justify-content: center; padding: 40px; margin: 0; }
        .container { background: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); width: 100%; max-width: 600px; }
        h2 { margin-top: 0; color: #333; text-align: center; border-bottom: 2px solid #f0f2f5; padding-bottom: 10px;}
        h3 { color: #555; font-size: 16px; margin-top: 25px; border-bottom: 1px solid #eee; padding-bottom: 5px;}
        .form-group { margin-bottom: 20px; }
        label { display: block; font-weight: bold; margin-bottom: 8px; color: #555; }
        input[type="text"], input[type="file"], select, textarea { width: 100%; padding: 12px; border: 1px solid #ccc; border-radius: 6px; box-sizing: border-box; font-size: 14px; }
        textarea { height: 80px; resize: vertical; }
        button { width: 100%; padding: 12px; background-color: #007bff; color: white; border: none; border-radius: 6px; font-size: 16px; font-weight: bold; cursor: pointer; transition: background 0.3s; }
        button:hover { background-color: #0056b3; }
        .btn-success { background-color: #28a745; margin-top: 10px; }
        .btn-success:hover { background-color: #218838; }
        .alert { padding: 12px; margin-top: 20px; border-radius: 6px; display: none; text-align: center; font-weight: bold;}
        .alert-success { background-color: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
        .info-text { font-size: 12px; color: #666; margin-top: 5px; display: block; }
    </style>
</head>
<body>
    <div class="container">
        <h2>Input Data Knowledge Base</h2>
        
        <!-- Form Input Manual -->
        <h3>1. Tambah Data Manual</h3>
        <form id="kategoriForm">
            <div class="form-group">
                <label for="nama">Nama Pengguna</label>
                <input type="text" id="nama" placeholder="Boleh dikosongkan (opsional)">
            </div>
            
            <div class="form-group">
                <label for="kategori">Kategori Layanan</label>
                <select id="kategori" required>
                    <option value="">-- Pilih Kategori --</option>
                    <option value="CRM">CRM (Informasi Umum & Hubungan Pelanggan)</option>
                    <option value="Sales">Sales (Pembelian Paket & Penawaran)</option>
                    <option value="Komplain">Komplain (Kendala & Pengaduan Teknis)</option>
                </select>
            </div>
            
            <div class="form-group">
                <label for="pertanyaan">Pertanyaan</label>
                <textarea id="pertanyaan" required placeholder="Tuliskan pertanyaan di sini..."></textarea>
            </div>

            <div class="form-group">
                <label for="jawaban">Jawaban</label>
                <textarea id="jawaban" required placeholder="Tuliskan jawaban dari pertanyaan di atas..."></textarea>
            </div>
            
            <button type="submit">Simpan Data Manual</button>
        </form>

        <!-- Form Upload CSV -->
        <h3>2. Upload via Excel (.xlsx)</h3>
        <form id="uploadForm">
            <div class="form-group">
                <label for="kategoriCsv">Kategori Layanan (Untuk semua data di file)</label>
                <select id="kategoriCsv" required>
                    <option value="">-- Pilih Kategori --</option>
                    <option value="CRM">CRM (Informasi Umum & Hubungan Pelanggan)</option>
                    <option value="Sales">Sales (Pembelian Paket & Penawaran)</option>
                    <option value="Komplain">Komplain (Kendala & Pengaduan Teknis)</option>
                </select>
            </div>
            <div class="form-group">
                <label for="fileXlsx">Pilih File Excel (.xlsx)</label>
                <input type="file" id="fileXlsx" accept=".xlsx, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" required>
                <span class="info-text">Format kolom di Excel: <b>A=Nama, B=Pertanyaan, C=Jawaban</b> (Baris pertama adalah header).</span>
            </div>
            <button type="submit" class="btn-success">Upload & Simpan Data</button>
        </form>

        <div id="status" class="alert alert-success"></div>
    </div>

    <script>
        // Handle Form Manual
        document.getElementById('kategoriForm').addEventListener('submit', async function(e) {
            e.preventDefault();
            const data = {
                nama: document.getElementById('nama').value,
                kategori: document.getElementById('kategori').value,
                pertanyaan: document.getElementById('pertanyaan').value,
                jawaban: document.getElementById('jawaban').value
            };
            try {
                const response = await fetch('/api/simpan', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(data)
                });
                if (response.ok) {
                    const statusEl = document.getElementById('status');
                    statusEl.style.display = 'block';
                    statusEl.innerText = 'Data manual berhasil disimpan ke JSON!';
                    document.getElementById('kategoriForm').reset();
                    setTimeout(() => { statusEl.style.display = 'none'; }, 4000);
                }
            } catch (error) { alert('Terjadi kesalahan saat menyimpan data.'); }
        });

        // Handle Upload CSV
        document.getElementById('uploadForm').addEventListener('submit', async function(e) {
            e.preventDefault();
            const fileInput = document.getElementById('fileXlsx');
            const kategoriCsv = document.getElementById('kategoriCsv').value;
            if (fileInput.files.length === 0) return;
            
            const formData = new FormData();
            formData.append('file', fileInput.files[0]);
            formData.append('kategori', kategoriCsv);

            try {
                const response = await fetch('/api/upload-xlsx', {
                    method: 'POST',
                    body: formData
                });
                if (response.ok) {
                    const resData = await response.json();
                    const statusEl = document.getElementById('status');
                    statusEl.style.display = 'block';
                    statusEl.innerText = `Berhasil! ${resData.added} baris data dari Excel ditambahkan ke JSON.`;
                    document.getElementById('uploadForm').reset();
                    setTimeout(() => { statusEl.style.display = 'none'; }, 4000);
                } else {
                    alert('Gagal mengupload file Excel. Pastikan format kolom dan file sudah benar.');
                }
            } catch (error) { alert('Terjadi kesalahan saat upload.'); }
        });
    </script>
</body>
</html>
"""

@app.get("/", response_class=HTMLResponse)
async def tampilkan_ui():
    """Menampilkan halaman HTML UI di browser"""
    return HTML_TEMPLATE

@app.post("/api/simpan")
async def simpan_data(data: LayananInput):
    """Menyimpan data dari UI ke dalam file JSON sementara"""
    records = load_json_records(JSON_FILE)
    waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    nama = data.nama if data.nama else "Anonim"
    
    new_record = {
        "id_tiket": f"TKT-{len(records)+1:04d}",
        "waktu": waktu,
        "nama": nama,
        "kategori": data.kategori,
        "pertanyaan": data.pertanyaan,
        "jawaban": data.jawaban,
        "status": "Terbuka"
    }
    records.append(new_record)
    with open(JSON_FILE, "w", encoding="utf-8") as f: json.dump(records, f, indent=4, ensure_ascii=False)
    
    return {"status": "success", "inserted_id": new_record["id_tiket"]}

@app.post("/api/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...), kategori: str = Form(...)):
    """Menerima file Excel (.xlsx) dari UI dan menyimpannya ke JSON (Support Emoji)"""
    records = load_json_records(JSON_FILE)
    content = await file.read()
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active

    data_to_insert = []
    waktu_sekarang = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Ambil semua baris
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) > 1:
        # Ambil header dan ubah ke huruf kecil
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        
        for row in rows[1:]:
            row_dict = {}
            for j in range(len(headers)):
                if headers[j] and j < len(row):
                    row_dict[headers[j]] = row[j] if row[j] else ""
                    
            # Cari nilai berdasarkan nama header
            pertanyaan = str(row_dict.get("pertanyaan", "")).strip()
            jawaban = str(row_dict.get("jawaban", "")).strip()
            nama = str(row_dict.get("nama", "Anonim")).strip()
            
            if not nama or nama == "None":
                nama = "Anonim"
    
            if pertanyaan and jawaban and pertanyaan != "None" and jawaban != "None":
                new_record = {
                    "id_tiket": f"TKT-{len(records)+1+len(data_to_insert):04d}",
                    "waktu": waktu_sekarang,
                    "nama": nama,
                    "kategori": kategori,
                    "pertanyaan": pertanyaan,
                    "jawaban": jawaban,
                    "status": "Terbuka"
                }
                data_to_insert.append(new_record)

    if data_to_insert:
        records.extend(data_to_insert)
        with open(JSON_FILE, "w", encoding="utf-8") as f: json.dump(records, f, indent=4, ensure_ascii=False)

    return {"status": "success", "added": len(data_to_insert)}

if __name__ == "__main__":
    print("Menjalankan UI di http://localhost:8080")
    uvicorn.run(app, host="0.0.0.0", port=8080)